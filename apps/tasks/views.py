import io
import json
from datetime import timedelta
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, UpdateView, DeleteView

from .forms import TaskForm
from .models import Task, ExecutionHistory
from apps.goals.models import Goal
from apps.ml.predictor import DurationPredictor
from apps.schedule.scheduler import TaskScheduler


class BoardView(LoginRequiredMixin, View):
    def get(self, request):
        tasks = Task.objects.filter(user=request.user).select_related('goal')
        columns = {
            'new': list(tasks.filter(status=Task.STATUS_NEW)),
            'planned': list(tasks.filter(status=Task.STATUS_PLANNED)),
            'in_progress': list(tasks.filter(status=Task.STATUS_IN_PROGRESS)),
            'done': list(tasks.filter(status=Task.STATUS_DONE)),
        }
        form = TaskForm()
        goals = Goal.objects.filter(user=request.user)
        return render(request, 'board/index.html', {
            'columns': columns,
            'form': form,
            'goals': goals,
        })


class BoardColumnsView(LoginRequiredMixin, View):
    """Returns rendered HTML for all kanban columns — used by chat to refresh board without reload."""
    def get(self, request):
        tasks = Task.objects.filter(user=request.user).select_related('goal')
        result = {}
        for status in [Task.STATUS_NEW, Task.STATUS_PLANNED, Task.STATUS_IN_PROGRESS, Task.STATUS_DONE]:
            task_list = list(tasks.filter(status=status))
            cards_html = ''.join(
                render_to_string('board/_task_card.html', {'task': t}, request=request)
                for t in task_list
            )
            if not task_list:
                cards_html = '<div class="text-muted text-center small py-3 empty-placeholder"><i class="bi bi-inbox"></i><br>Нет задач</div>'
            result[status] = {'html': cards_html, 'count': len(task_list)}
        return JsonResponse(result)


class TaskCreateView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        form = TaskForm(data)
        if form.is_valid():
            task = form.save(commit=False)
            task.user = request.user
            if task.status == Task.STATUS_DONE and not task.completed_at:
                task.completed_at = timezone.now()
            task.save()

            if task.status == Task.STATUS_DONE:
                correction_factor = None
                if task.estimated_duration and task.actual_duration:
                    correction_factor = task.actual_duration / task.estimated_duration
                ExecutionHistory.objects.update_or_create(
                    task=task,
                    defaults={
                        'user': request.user,
                        'category': task.category,
                        'priority': task.priority,
                        'estimated_duration': task.estimated_duration or None,
                        'actual_duration': task.actual_duration or None,
                        'correction_factor': correction_factor,
                        'completed_at': task.completed_at,
                    }
                )
            else:
                try:
                    TaskScheduler(request.user).schedule_task(task)
                    task.refresh_from_db()
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).error('schedule_task failed: %s', e)

            task_html = render_to_string('board/_task_card.html', {'task': task}, request=request)
            return JsonResponse({'success': True, 'task_html': task_html, 'task_id': task.id})
        return JsonResponse({'success': False, 'errors': form.errors})


class TaskUpdateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = get_object_or_404(Task, pk=pk, user=request.user)
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        form = TaskForm(data, instance=task)
        if form.is_valid():
            task = form.save()
            if task.estimated_duration:
                TaskScheduler(request.user).schedule_task(task)
                task.refresh_from_db()
            task_html = render_to_string('board/_task_card.html', {'task': task}, request=request)
            return JsonResponse({'success': True, 'task_html': task_html})
        return JsonResponse({'success': False, 'errors': form.errors})


class TaskDeleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        task = get_object_or_404(Task, pk=pk, user=request.user)
        task.delete()
        return JsonResponse({'success': True})


class TaskStatusView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        task_id = data.get('task_id')
        new_status = data.get('status')

        if not task_id or not new_status:
            return JsonResponse({'success': False, 'error': 'Missing task_id or status'})

        valid_statuses = [s[0] for s in Task.STATUS_CHOICES]
        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'error': 'Invalid status'})

        task = get_object_or_404(Task, pk=task_id, user=request.user)
        task.status = new_status
        task.save(update_fields=['status', 'updated_at'])

        scheduled_start = None
        scheduled_end = None
        is_overdue_risk = task.is_overdue_risk

        if new_status == Task.STATUS_PLANNED and task.estimated_duration:
            scheduler = TaskScheduler(request.user)
            result = scheduler.schedule_task(task)
            if result:
                scheduled_start, scheduled_end, is_overdue_risk = result
                scheduled_start = scheduled_start.isoformat() if scheduled_start else None
                scheduled_end = scheduled_end.isoformat() if scheduled_end else None

        return JsonResponse({
            'success': True,
            'scheduled_start': scheduled_start,
            'scheduled_end': scheduled_end,
            'is_overdue_risk': is_overdue_risk,
        })


class TaskCompleteView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        task_id = data.get('task_id')
        actual_duration = data.get('actual_duration')

        task = get_object_or_404(Task, pk=task_id, user=request.user)

        task.status = Task.STATUS_DONE
        task.completed_at = timezone.now()
        task.is_overdue_risk = False
        if actual_duration:
            try:
                task.actual_duration = int(actual_duration)
            except (ValueError, TypeError):
                pass
        task.save()

        correction_factor = None
        if task.estimated_duration and task.actual_duration:
            correction_factor = task.actual_duration / task.estimated_duration

        ExecutionHistory.objects.update_or_create(
            task=task,
            defaults={
                'user': request.user,
                'category': task.category,
                'priority': task.priority,
                'estimated_duration': task.estimated_duration or None,
                'actual_duration': task.actual_duration or None,
                'correction_factor': correction_factor,
                'completed_at': task.completed_at,
            }
        )

        if task.goal:
            task.goal.check_achieved()

        return JsonResponse({'success': True})


class TaskAssignGoalView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        task_id = data.get('task_id')
        goal_id = data.get('goal_id')
        goal_title = data.get('goal_title', '').strip()

        task = get_object_or_404(Task, pk=task_id, user=request.user)

        if goal_id:
            goal = get_object_or_404(Goal, pk=goal_id, user=request.user)
        elif goal_title:
            goal, _ = Goal.objects.get_or_create(
                user=request.user,
                title=goal_title,
                defaults={'description': ''}
            )
        else:
            return JsonResponse({'success': False, 'error': 'No goal specified'})

        task.goal = goal
        task.save(update_fields=['goal', 'updated_at'])

        return JsonResponse({'success': True, 'goal_title': goal.title, 'goal_id': goal.id})


class TaskPredictTimeView(LoginRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = request.POST.dict()

        category = data.get('category', '')
        priority = data.get('priority', 'medium')
        try:
            estimated_duration = int(data.get('estimated_duration', 0))
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid estimated_duration'})

        if not category or not estimated_duration:
            return JsonResponse({'success': False, 'error': 'Missing category or estimated_duration'})

        predictor = DurationPredictor(request.user)
        meta = predictor.predict_with_meta(category, estimated_duration, priority)

        model_labels = {
            'RandomForest': 'случайный лес',
            'LinearRegression': 'линейная регрессия',
            'CorrectionFactor': 'поправочный коэффициент',
        }
        if meta['has_enough_data']:
            model_label = model_labels.get(meta['model_name'], meta['model_name'])
            message = (
                f"Модель: {model_label} · "
                f"{meta['n_samples']} выполненных задач · "
                f"коэффициент {meta['factor']:.2f}×"
            )
        else:
            message = f"Недостаточно данных (есть {meta['n_samples']}, нужно минимум 10)."

        return JsonResponse({
            'success': True,
            'predicted_duration': meta['predicted_duration'],
            'factor': meta['factor'],
            'has_enough_data': meta['has_enough_data'],
            'model_name': meta['model_name'],
            'n_samples': meta['n_samples'],
            'feature_importances': meta['feature_importances'],
            'message': message,
        })


class TaskRecommendView(LoginRequiredMixin, View):
    def get(self, request):
        now = timezone.now()

        # Priority 1: overdue high-priority tasks
        task = Task.objects.filter(
            user=request.user,
            status__in=[Task.STATUS_NEW, Task.STATUS_PLANNED, Task.STATUS_IN_PROGRESS],
            priority=Task.PRIORITY_HIGH,
            deadline__lt=now,
        ).first()
        if task:
            return JsonResponse({
                'task_id': task.id,
                'title': task.title,
                'reason': 'Просрочена высокоприоритетная задача!',
            })

        # Priority 2: tasks scheduled for now
        task = Task.objects.filter(
            user=request.user,
            status=Task.STATUS_PLANNED,
            scheduled_start__lte=now,
            scheduled_end__gte=now,
        ).first()
        if task:
            return JsonResponse({
                'task_id': task.id,
                'title': task.title,
                'reason': 'Сейчас запланировано время для этой задачи.',
            })

        # Priority 3: in progress tasks
        task = Task.objects.filter(
            user=request.user,
            status=Task.STATUS_IN_PROGRESS,
        ).first()
        if task:
            return JsonResponse({
                'task_id': task.id,
                'title': task.title,
                'reason': 'Задача уже в работе — продолжайте!',
            })

        # Priority 4: high priority new tasks with nearest deadline
        task = Task.objects.filter(
            user=request.user,
            status=Task.STATUS_NEW,
            priority=Task.PRIORITY_HIGH,
        ).order_by('deadline').first()
        if task:
            return JsonResponse({
                'task_id': task.id,
                'title': task.title,
                'reason': 'Высокоприоритетная задача требует внимания.',
            })

        # Priority 5: any planned task
        task = Task.objects.filter(
            user=request.user,
            status__in=[Task.STATUS_NEW, Task.STATUS_PLANNED],
        ).order_by('deadline').first()
        if task:
            return JsonResponse({
                'task_id': task.id,
                'title': task.title,
                'reason': 'Ближайшая задача по сроку.',
            })

        return JsonResponse({'task_id': None, 'message': 'Все задачи выполнены! Отличная работа.'})


class TaskDetailView(LoginRequiredMixin, View):
    def get(self, request, pk):
        task = get_object_or_404(Task, pk=pk, user=request.user)
        return JsonResponse({
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'category': task.category,
            'priority': task.priority,
            'status': task.status,
            'deadline': task.deadline.isoformat() if task.deadline else None,
            'estimated_duration': task.estimated_duration,
            'actual_duration': task.actual_duration,
            'scheduled_start': task.scheduled_start.isoformat() if task.scheduled_start else None,
            'scheduled_end': task.scheduled_end.isoformat() if task.scheduled_end else None,
            'goal_id': task.goal_id,
            'goal_title': task.goal.title if task.goal else None,
            'is_overdue_risk': task.is_overdue_risk,
            'created_at': task.created_at.isoformat(),
            'priority_color': task.priority_color,
            'status_label': task.status_label,
        })



class TaskRescheduleView(LoginRequiredMixin, View):
    def post(self, request, pk):
        from django.utils.dateparse import parse_datetime
        task = get_object_or_404(Task, pk=pk, user=request.user)
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception):
            data = {}

        scheduled_start_str = data.get('scheduled_start')
        if not scheduled_start_str:
            return JsonResponse({'success': False, 'error': 'scheduled_start обязателен'})

        scheduled_start = parse_datetime(scheduled_start_str)
        if not scheduled_start:
            return JsonResponse({'success': False, 'error': 'Неверный формат даты'})

        if timezone.is_naive(scheduled_start):
            scheduled_start = timezone.make_aware(scheduled_start)


        duration = task.estimated_duration or 60
        scheduled_end = scheduled_start + timedelta(minutes=duration)

        task.scheduled_start = scheduled_start
        task.scheduled_end = scheduled_end
        if task.status == Task.STATUS_NEW:
            task.status = Task.STATUS_PLANNED
        task.save(update_fields=['scheduled_start', 'scheduled_end', 'status', 'updated_at'])

        return JsonResponse({
            'success': True,
            'scheduled_start': task.scheduled_start.isoformat(),
            'scheduled_end': task.scheduled_end.isoformat(),
        })


class TaskExportExcelView(LoginRequiredMixin, View):

    STATUS_LABELS   = {'new': 'Новая', 'planned': 'Запланирована', 'in_progress': 'В работе', 'done': 'Выполнена'}
    PRIORITY_LABELS = {'high': 'Высокий', 'medium': 'Средний', 'low': 'Низкий'}
    CATEGORY_LABELS = {'study': 'Учёба', 'work': 'Работа', 'household': 'Быт',
                       'health': 'Здоровье', 'personal': 'Личное', 'other': 'Другое'}

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
        from openpyxl.utils import get_column_letter

        tasks = list(Task.objects.filter(user=request.user).order_by('status', '-created_at'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Задачи'
        ws.sheet_view.showGridLines = False

        headers = ['Название', 'Описание', 'Категория', 'Приоритет', 'Статус',
                   'Дедлайн', 'Оценка (мин)', 'Факт (мин)', 'Запланировано', 'Создана', 'Выполнена']
        col_widths = [32, 42, 14, 13, 17, 18, 14, 12, 20, 18, 18]

        # Стили
        header_fill = PatternFill('solid', fgColor='1A3C5E')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thick = Side(style='medium', color='1A3C5E')
        thin  = Side(style='thin',   color='D0D7E0')
        header_border = Border(left=thick, right=thick, top=thick, bottom=thick)
        cell_border   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)

        row_even = PatternFill('solid', fgColor='EEF3FA')
        row_odd  = PatternFill('solid', fgColor='FFFFFF')

        priority_colors = {'high': 'C0392B', 'medium': 'D4AC0D', 'low': '1E8449'}
        status_colors   = {
            'new':         '7F8C8D',
            'planned':     '2471A3',
            'in_progress': 'D68910',
            'done':        '1E8449',
        }

        # Заголовки
        for col, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = header_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 26

        def fmt_dt(dt):
            if not dt:
                return '—'
            return timezone.localtime(dt).strftime('%d.%m.%Y %H:%M')

        for row_idx, task in enumerate(tasks, 2):
            is_even = row_idx % 2 == 0
            row_fill = row_even if is_even else row_odd

            values = [
                task.title,
                task.description or '',
                self.CATEGORY_LABELS.get(task.category, task.category),
                self.PRIORITY_LABELS.get(task.priority, task.priority),
                self.STATUS_LABELS.get(task.status, task.status),
                fmt_dt(task.deadline),
                task.estimated_duration or '—',
                task.actual_duration or '—',
                fmt_dt(task.scheduled_start),
                fmt_dt(task.created_at),
                fmt_dt(task.completed_at),
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border    = cell_border
                cell.alignment = Alignment(vertical='center', wrap_text=(col == 2))

                # Название — жирное
                if col == 1:
                    cell.font = Font(name='Calibri', bold=True, size=10)
                    cell.fill = row_fill
                # Приоритет — цветной текст
                elif col == 4:
                    color = priority_colors.get(task.priority, '2C3E50')
                    cell.font  = Font(name='Calibri', bold=True, color=color, size=10)
                    cell.fill  = row_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Статус — цветной текст
                elif col == 5:
                    color = status_colors.get(task.status, '2C3E50')
                    cell.font  = Font(name='Calibri', bold=True, color=color, size=10)
                    cell.fill  = row_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                # Числа по центру
                elif col in (7, 8):
                    cell.font      = Font(name='Calibri', size=10)
                    cell.fill      = row_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='Calibri', size=10)
                    cell.fill = row_fill

            ws.row_dimensions[row_idx].height = 20

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:K{len(tasks) + 1}'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = timezone.localtime(timezone.now()).strftime('%Y-%m-%d')
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="tasks_{date_str}.xlsx"'
        return response

